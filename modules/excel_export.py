"""
Excel Export — implements the xlsx-manipulation / excel-automation skills.

Builds styled openpyxl workbooks in-memory (BytesIO) for download buttons:
  • tracked_portfolio_xlsx(tp, analysis, plan) — Tracker report
  • portfolio_xlsx(df, summary)                — uploaded Portfolio report
"""
import io
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HDR_FILL = PatternFill("solid", fgColor="1F2937")
_HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
_TITLE_FONT = Font(bold=True, size=14)
_POS_FONT = Font(color="16A34A")
_NEG_FONT = Font(color="DC2626")
_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill, c.font, c.border = _HDR_FILL, _HDR_FONT, _BORDER
        c.alignment = Alignment(horizontal="center")


def _autofit(ws, widths: dict[int, int] | None = None) -> None:
    for col_cells in ws.columns:
        col = col_cells[0].column
        maxlen = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col)].width = min(40, max(10, maxlen + 2))
    for col, w in (widths or {}).items():
        ws.column_dimensions[get_column_letter(col)].width = w


def tracked_portfolio_xlsx(tp: dict, analysis: dict, plan: dict) -> bytes:
    """Full Tracker report: Summary, Holdings, Rebalance Plan, History."""
    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Tracked Portfolio Report — {tp.get('name', '')}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Generated {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(color="6B7280", size=9)

    rows = [
        ("Created",        tp.get("created_at", "")),
        ("Thesis",         tp.get("thesis", "")),
        ("Risk Level",     tp.get("risk_level", "")),
        ("Initial Capital", tp.get("capital", 0)),
        ("Current Value",  round(analysis.get("total_value", 0), 2)),
        ("Total Return %", round(analysis.get("total_return", 0), 2)),
        ("Market Regime",  analysis.get("regime", {}).get("regime", "")),
        ("Fired Triggers", analysis.get("n_triggers", 0)),
        ("Warnings",       " | ".join(analysis.get("warnings", []))),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    _autofit(ws, {2: 60})

    # ── Holdings sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Holdings")
    headers = ["Symbol", "Name", "Role", "Sector", "Target %", "Current %", "Drift pp",
               "Entry $", "Price $", "Value $", "Since Entry %", "1M %", "3M %",
               "RSI", "Fwd P/E", "Insider MSPR", "Score", "Signal", "Triggers"]
    _write_header(ws2, 1, headers)
    for r_i, r in enumerate(sorted(analysis.get("positions", []),
                                   key=lambda x: -x.get("current_weight", 0)), start=2):
        vals = [
            r["symbol"], r.get("name", ""), r.get("role", ""), r.get("sector", ""),
            round(r["target_weight"], 2), round(r["current_weight"], 2), round(r["drift"], 2),
            round(r["entry_price"], 2), round(r["price"], 2), round(r["value"], 2),
            round(r["ret_since_entry"], 2),
            round((r.get("r1m") or 0) * 100, 2) if r.get("r1m") is not None else None,
            round((r.get("r3m") or 0) * 100, 2) if r.get("r3m") is not None else None,
            round(r["rsi"], 0) if r.get("rsi") is not None else None,
            round(r["fwd_pe"], 1) if r.get("fwd_pe") else None,
            r.get("mspr"),
            r.get("score"),
            r.get("score_label", ""),
            "; ".join(msg for _, msg in r.get("triggers", [])),
        ]
        for c_i, v in enumerate(vals, start=1):
            cell = ws2.cell(row=r_i, column=c_i, value=v)
            cell.border = _BORDER
        # color drift + since-entry
        for col_idx, val in ((7, vals[6]), (11, vals[10])):
            if isinstance(val, (int, float)):
                ws2.cell(row=r_i, column=col_idx).font = _POS_FONT if val >= 0 else _NEG_FONT
    ws2.freeze_panes = "A2"
    _autofit(ws2, {19: 50})

    # ── Rebalance Plan sheet ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Rebalance Plan")
    _write_header(ws3, 1, ["Symbol", "Action", "Amount $", "Drift pp"])
    for r_i, a in enumerate(plan.get("actions", []), start=2):
        ws3.cell(row=r_i, column=1, value=a["symbol"]).border = _BORDER
        act = ws3.cell(row=r_i, column=2, value=a["action"])
        act.font = _POS_FONT if a["action"] == "BUY" else _NEG_FONT
        act.border = _BORDER
        ws3.cell(row=r_i, column=3, value=abs(a["amount_usd"])).border = _BORDER
        ws3.cell(row=r_i, column=4, value=a["drift"]).border = _BORDER
    base = len(plan.get("actions", [])) + 3
    ws3.cell(row=base,     column=1, value="Turnover $").font = Font(bold=True)
    ws3.cell(row=base,     column=2, value=plan.get("turnover_usd", 0))
    ws3.cell(row=base + 1, column=1, value="Est. Cost $").font = Font(bold=True)
    ws3.cell(row=base + 1, column=2, value=plan.get("est_cost_usd", 0))
    ws3.cell(row=base + 2, column=1, value="Worth It").font = Font(bold=True)
    ws3.cell(row=base + 2, column=2, value="Yes" if plan.get("worth_it") else "Marginal")
    _autofit(ws3)

    # ── Value History sheet ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Value History")
    _write_header(ws4, 1, ["Date", "Value $"])
    for r_i, h in enumerate(tp.get("value_history", []), start=2):
        ws4.cell(row=r_i, column=1, value=h.get("date", "")).border = _BORDER
        ws4.cell(row=r_i, column=2, value=h.get("value", 0)).border = _BORDER
    _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def portfolio_xlsx(df: pd.DataFrame, summary: dict, title: str = "Portfolio") -> bytes:
    """Uploaded-portfolio report: Summary + Holdings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"{title} Report"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Generated {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(color="6B7280", size=9)

    srows = []
    for key in ("total_value", "total_cost", "total_pnl", "total_pnl_pct",
                "total_value_nis", "total_cost_nis", "total_pnl_nis", "n_positions"):
        if key in summary:
            v = summary[key]
            if key == "total_pnl_pct":
                v = round(v * 100, 2)
            elif isinstance(v, float):
                v = round(v, 2)
            srows.append((key.replace("_", " ").title(), v))
    for i, (k, v) in enumerate(srows, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    sector_exp = summary.get("sector_exposure")
    if sector_exp is not None and len(sector_exp):
        start = len(srows) + 6
        ws.cell(row=start, column=1, value="Sector Exposure %").font = Font(bold=True)
        for i, (sec, w) in enumerate(sector_exp.items(), start=start + 1):
            ws.cell(row=i, column=1, value=str(sec))
            ws.cell(row=i, column=2, value=round(float(w), 1))
    _autofit(ws)

    ws2 = wb.create_sheet("Holdings")
    cols = [c for c in df.columns if c != "Price Stale"]
    _write_header(ws2, 1, cols)
    for r_i, (_, row) in enumerate(df.iterrows(), start=2):
        for c_i, col in enumerate(cols, start=1):
            v = row[col]
            if isinstance(v, float):
                v = round(v, 4)
            cell = ws2.cell(row=r_i, column=c_i, value=v if pd.notna(v) else None)
            cell.border = _BORDER
            if col in ("P&L ($)", "P&L (%)", "P&L (NIS)") and isinstance(v, (int, float)):
                cell.font = _POS_FONT if v >= 0 else _NEG_FONT
    ws2.freeze_panes = "A2"
    _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
